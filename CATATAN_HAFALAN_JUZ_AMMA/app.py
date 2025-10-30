import streamlit as st
import pandas as pd
import json
from datetime import datetime
import os
import plotly.express as px

# Import data master dan fungsi pembantu dari file juz_amma_data.py
from juz_amma_data import (
    JUZ_AMMA_MAP,
    SURAH_NAMES,
    TOTAL_AYAT_JUZ_AMMA,
    initialize_database,
    create_initial_data_structure,
    calculate_lulus_count,
)

# =============================
# KONFIGURASI APLIKASI / FILE
# =============================
#DB_FILE = "data_hafalan.csv"          # database utama murid + status hafalannya
#GURU_FILE = "guru_list.csv"          # daftar guru pencatat (dropdown)
#LOG_FILE = "log_hafalan.csv"         # riwayat transaksi setoran hafalan

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, "data_hafalan.csv")
GURU_FILE = os.path.join(BASE_DIR, "guru_list.csv")
LOG_FILE = os.path.join(BASE_DIR, "log_hafalan.csv")
logo_path = os.path.join(BASE_DIR, "logo.png")

if os.path.exists(logo_path):
    st.sidebar.image(logo_path, width=120)
else:
    st.sidebar.markdown("**SMP Negeri 9 Banjar**")

# Pastikan file CSV penting tersedia
for filename, header in [
    ("data_hafalan.csv", "ID_Murid,Nama_Murid,Kelas,Status_Hafalan,Total_Ayat_Lulus,Update_Terakhir,Guru_Pencatat"),
    ("guru_list.csv", "Nama_Guru"),
    ("log_hafalan.csv", "Timestamp,ID_Murid,Nama_Murid,Kelas,Surah,Ayat_Dari,Ayat_Sampai,Status,Guru_Pencatat"),
]:
    if not os.path.exists(filename):
        with open(filename, "w", encoding="utf-8") as f:
            f.write(header + "\\n")
        st.warning(f"File {filename} tidak ditemukan, dibuat otomatis.")

st.set_page_config(
    page_title="Pencatatan Hafalan Juz Amma",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Load database awal ke session_state jika belum ada
if "df" not in st.session_state:
    st.session_state.df = initialize_database(DB_FILE)

# =============================
# FUNGSI UTILITAS / DATA
# =============================

def load_guru_list(csv_path: str = GURU_FILE):
    """
    Membaca daftar guru dari file CSV.
    Jika file tidak ditemukan atau formatnya tidak sesuai, buat file contoh otomatis.
    """
    default_guru = [
        "Agus Sugiharto Sapari, S.Pd.",
        "Siti Maryam, S.Pd.",
        "Rahmat Hidayat, S.Pd.I.",
        "Nisa Khairun, S.Pd.",
    ]

    if not os.path.exists(csv_path):
        st.warning(f"File '{csv_path}' tidak ditemukan. Membuat file contoh otomatis.")
        pd.DataFrame({"Nama_Guru": default_guru}).to_csv(csv_path, index=False)
        return ["Pilih Guru"] + default_guru

    try:
        # Tambahkan opsi engine dan delimiter fallback
        try:
            df_guru = pd.read_csv(csv_path, sep=",", engine="python")
        except pd.errors.ParserError:
            df_guru = pd.read_csv(csv_path, sep=";", engine="python")
        except Exception:
            # fallback terakhir: coba baca sebagai satu kolom
            df_guru = pd.read_csv(csv_path, header=None, names=["Nama_Guru"])

        if "Nama_Guru" not in df_guru.columns:
            st.warning(f"File '{csv_path}' tidak memiliki kolom 'Nama_Guru'. Menggunakan daftar default.")
            return ["Pilih Guru"] + default_guru

        guru_list = ["Pilih Guru"] + df_guru["Nama_Guru"].dropna().astype(str).tolist()
        return guru_list

    except Exception as e:
        st.error(f"Gagal membaca '{csv_path}': {e}")
        return ["Pilih Guru"] + default_guru


def ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pastikan kolom penting selalu ada di dataframe meski file lama.
    """
    if "Guru_Pencatat" not in df.columns:
        df["Guru_Pencatat"] = ""
    if "NIS" not in df.columns:
        df["NIS"] = ""
    if "Total_Ayat_Lulus" not in df.columns:
        df["Total_Ayat_Lulus"] = 0
    if "Update_Terakhir" not in df.columns:
        df["Update_Terakhir"] = ""
    return df


def save_data(df: pd.DataFrame):
    """
    Simpan df terbaru ke CSV utama dan update session_state.
    """
    df = ensure_columns(df)
    df.to_csv(DB_FILE, index=False)
    st.session_state.df = df


def add_new_student(name, kelas, nis=""):
    """
    Tambah murid baru manual via sidebar.
    """
    df = ensure_columns(st.session_state.df.copy())
    next_id = df["ID_Murid"].max() + 1 if not df.empty else 1001

    new_data = {
        "ID_Murid": next_id,
        "Nama_Murid": name,
        "NIS": nis,
        "Kelas": kelas,
        "Status_Hafalan": create_initial_data_structure(),
        "Total_Ayat_Lulus": 0,
        "Update_Terakhir": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Guru_Pencatat": "",
    }

    new_df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)
    save_data(new_df)
    st.success(f"Murid **{name}** (ID: {next_id}) berhasil ditambahkan ke kelas **{kelas}**.")


def import_students_from_csv(uploaded_file):
    """
    Impor massal murid dari CSV (pemisah ;). Wajib kolom: Nama_Murid, Kelas. Opsional: NIS.
    """
    try:
        new_students_df = pd.read_csv(uploaded_file, sep=';')

        REQUIRED_COLS = ['Nama_Murid', 'Kelas']
        if not all(col in new_students_df.columns for col in REQUIRED_COLS):
            st.error("File CSV harus memiliki kolom wajib: Nama_Murid, Kelas")
            st.info("Pastikan CSV menggunakan pemisah ';'")
            return

        new_students_df = new_students_df.dropna(subset=REQUIRED_COLS)
        if new_students_df.empty:
            st.warning("Tidak ada baris murid valid di CSV.")
            return

        df = ensure_columns(st.session_state.df.copy())
        current_max_id = df['ID_Murid'].max() if not df.empty else 1000
        num_new = len(new_students_df)
        new_ids = range(int(current_max_id) + 1, int(current_max_id) + 1 + num_new)
        new_students_df['ID_Murid'] = new_ids

        new_students_df['Status_Hafalan'] = [create_initial_data_structure() for _ in range(num_new)]
        new_students_df['Total_Ayat_Lulus'] = 0
        new_students_df['Update_Terakhir'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_students_df['Guru_Pencatat'] = ""
        if 'NIS' not in new_students_df.columns:
            new_students_df['NIS'] = ""

        combined_df = pd.concat([df, new_students_df], ignore_index=True)
        save_data(combined_df)

        st.success(f"{num_new} murid berhasil diimpor!")
        st.info("Cek menu lain untuk melihat data baru.")

    except Exception as e:
        st.error(f"Terjadi kesalahan saat memproses file: {e}")
        st.warning("Pastikan file CSV valid dan menggunakan ';' sebagai pemisah kolom.")


def log_setoran(student_row, surah, start_ayat, end_ayat, status_code, guru_pencatat):
    """
    Mencatat transaksi setoran hafalan ke file LOG_FILE.
    Tiap klik Simpan Catatan -> 1 baris baru masuk log.
    """
    status_label = "Lulus" if status_code == 1 else "Mengulang"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    log_data = {
        "Timestamp": now,
        "ID_Murid": student_row["ID_Murid"],
        "Nama_Murid": student_row["Nama_Murid"],
        "Kelas": student_row["Kelas"],
        "Surah": surah,
        "Ayat_Dari": start_ayat,
        "Ayat_Sampai": end_ayat,
        "Status": status_label,
        "Guru_Pencatat": guru_pencatat,
    }

    write_header = not os.path.exists(LOG_FILE)
    pd.DataFrame([log_data]).to_csv(LOG_FILE, mode="a", index=False, header=write_header)


def update_hafalan_status(
    df: pd.DataFrame,
    student_id: int,
    surah: str,
    start_ayat: int,
    end_ayat: int,
    status_code: int,
    guru_pencatat: str,
):
    """
    Update status hafalan ayat tertentu untuk murid.
    Sekaligus catat log transaksi setoran guru ke LOG_FILE.
    """
    max_ayat = JUZ_AMMA_MAP.get(surah)
    if not max_ayat:
        st.error("Nama surah tidak valid.")
        return df

    if not (1 <= start_ayat <= max_ayat and 1 <= end_ayat <= max_ayat and start_ayat <= end_ayat):
        st.error(f"Rentang ayat tidak valid. Surah {surah} hanya punya ayat 1 sampai {max_ayat}.")
        return df

    idx_list = df[df['ID_Murid'] == student_id].index
    if idx_list.empty:
        st.error("Murid tidak ditemukan.")
        return df
    idx = idx_list[0]

    # Ambil dan parse status hafalan (JSON → dict)
    try:
        status_dict = json.loads(df.loc[idx, 'Status_Hafalan'])
    except Exception:
        st.error("Format Status_Hafalan rusak untuk murid ini.")
        return df

    # Update per ayat
    ayat_list = status_dict.get(surah, [])
    for i in range(start_ayat - 1, end_ayat):  # 0-index
        if i < len(ayat_list):
            ayat_list[i] = status_code
    status_dict[surah] = ayat_list

    # Simpan balik (dict → JSON string)
    new_status_json = json.dumps(status_dict)
    df.loc[idx, 'Status_Hafalan'] = new_status_json
    df.loc[idx, 'Total_Ayat_Lulus'] = calculate_lulus_count(new_status_json)
    df.loc[idx, 'Update_Terakhir'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df.loc[idx, 'Guru_Pencatat'] = guru_pencatat

    # Catat log transaksi setoran
    student_row = df.loc[df['ID_Murid'] == student_id].iloc[0]
    log_setoran(student_row, surah, start_ayat, end_ayat, status_code, guru_pencatat)

    # Simpan database utama
    save_data(df)

    st.success(
        f"Berhasil mencatat setoran {surah} ayat {start_ayat}-{end_ayat} sebagai "
        + ("LULUS" if status_code == 1 else "MENGULANG")
        + f". Dicatat oleh {guru_pencatat}."
    )

    return df


def delete_student(df, student_id, student_name):
    """
    Hapus murid dari database utama.
    """
    initial_len = len(df)
    new_df = df[df['ID_Murid'] != student_id].copy()

    if len(new_df) < initial_len:
        save_data(new_df)
        st.success(f"Murid **{student_name}** (ID: {student_id}) berhasil dihapus dari database.")
    else:
        st.error(f"Gagal menghapus. Murid dengan ID {student_id} tidak ditemukan.")
    return new_df

# =============================
# HALAMAN: INPUT SETORAN / PENCATATAN HAFALAN
# =============================

def page_pencatatan_hafalan(df, selected_class, selected_guru):
    st.header("📝 Input Setoran Hafalan per Murid")

    if selected_class == "Pilih Kelas":
        st.warning("Mohon pilih kelas di sidebar terlebih dahulu.")
        return

    # Filter murid per kelas
    class_df = df[df['Kelas'] == selected_class]

    student_map = {
        f"{row['Nama_Murid']} (ID: {row['ID_Murid']})": row['ID_Murid']
        for _, row in class_df.iterrows()
    }
    student_display_list = ['Pilih Murid'] + list(student_map.keys())

    selected_student_display = st.selectbox("Pilih Murid", student_display_list)
    if selected_student_display == 'Pilih Murid':
        return

    selected_student_id = student_map[selected_student_display]
    student_row = df[df['ID_Murid'] == selected_student_id].iloc[0]

    st.subheader(f"Murid: {student_row['Nama_Murid']}")

    progress_percent = int(
        (student_row['Total_Ayat_Lulus'] / TOTAL_AYAT_JUZ_AMMA) * 100
        if TOTAL_AYAT_JUZ_AMMA > 0 else 0
    )
    st.info(
        f"Total Ayat Lulus: {student_row['Total_Ayat_Lulus']} dari {TOTAL_AYAT_JUZ_AMMA} ayat.\n"
        f"Progres: {progress_percent}%.\n"
        f"Terakhir Diperbarui: {student_row['Update_Terakhir']}\n"
        f"Dicatat oleh: {student_row.get('Guru_Pencatat', '')}"
    )

    st.markdown("---")
    st.subheader("Riwayat Status Ayat per Surah")

    surah_to_setor = st.selectbox("Surah", SURAH_NAMES)
    max_ayat_current = JUZ_AMMA_MAP.get(surah_to_setor, 1)

    # tampilkan status per ayat surah yg dipilih
    try:
        status_dict_full = json.loads(student_row['Status_Hafalan'])
        ayat_list = status_dict_full.get(surah_to_setor, [])

        STATUS_LABELS = {
            0: "⚫ Belum",
            1: "🟢 Lulus",
            2: "🟠 Mengulang",
        }

        st.markdown(f"**Riwayat Status Ayat Surah {surah_to_setor} (total {max_ayat_current} ayat):**")

        num_columns = 5
        cols = st.columns(num_columns)
        for i, status_val in enumerate(ayat_list):
            ayat_num = i + 1
            col_index = i % num_columns
            label = STATUS_LABELS.get(status_val, "❓ Error")
            cols[col_index].markdown(f"**Ayat {ayat_num}**: {label}")
    except Exception as e:
        st.error(f"Gagal memuat riwayat hafalan: {e}")

    st.markdown("---")
    st.subheader("Formulir Setoran Baru")

    col3, col4 = st.columns(2)
    with col3:
        start_ayat = st.number_input(
            "Dari Ayat Ke-",
            min_value=1,
            max_value=max_ayat_current,
            value=1,
            key="start_ayat_input",
        )
    with col4:
        end_ayat = st.number_input(
            "Sampai Ayat Ke-",
            min_value=start_ayat,
            max_value=max_ayat_current,
            value=start_ayat,
            key="end_ayat_input",
        )

    setoran_status = st.radio(
        "Hasil Setoran:",
        options=["Lulus", "Mengulang"],
        index=0,
        horizontal=True,
    )
    status_code = 1 if setoran_status == "Lulus" else 2

    simpan_clicked = st.button("✅ Simpan Catatan")
    if simpan_clicked:
        if selected_guru == "Pilih Guru":
            st.warning("Pilih nama guru pencatat di sidebar terlebih dahulu.")
        else:
            update_hafalan_status(
                df.copy(),
                selected_student_id,
                surah_to_setor,
                start_ayat,
                end_ayat,
                status_code,
                selected_guru,
            )
            st.rerun()

# =============================
# HALAMAN: REKAP PER SURAH PER KELAS
# =============================

def build_rekap_per_surah(df, selected_class):
    class_df = df[df['Kelas'] == selected_class]
    hasil = []

    for surah, total_ayat in JUZ_AMMA_MAP.items():
        lulus_total = 0
        mengulang_total = 0
        belum_total = 0

        for _, row in class_df.iterrows():
            status_dict = json.loads(row['Status_Hafalan'])
            ayat_list = status_dict.get(surah, [])
            lulus_total += ayat_list.count(1)
            mengulang_total += ayat_list.count(2)
            belum_total += ayat_list.count(0)

        denom = len(class_df) * total_ayat if len(class_df) > 0 else 1
        persen = round((lulus_total / denom) * 100, 2) if denom > 0 else 0

        hasil.append({
            'Surah': surah,
            'Lulus': lulus_total,
            'Mengulang': mengulang_total,
            'Belum': belum_total,
            'Persentase Lulus (%)': persen,
        })

    return pd.DataFrame(hasil)


def page_rekap_per_surah(df, selected_class):
    st.header("📘 Rekap Hafalan per Surah (per Kelas)")

    if selected_class == "Pilih Kelas":
        st.info("Pilih kelas di sidebar untuk melihat rekap per surah.")
        return

    rekap_df = build_rekap_per_surah(df, selected_class)

    st.subheader(f"Rekap Kelas {selected_class}")
    st.dataframe(rekap_df, use_container_width=True)

    fig = px.bar(
        rekap_df,
        x='Surah',
        y='Persentase Lulus (%)',
        color='Persentase Lulus (%)',
        title=f"Persentase Ayat Lulus per Surah - Kelas {selected_class}",
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.subheader("📤 Unduh Rekap")

    csv_bytes = rekap_df.to_csv(index=False).encode('utf-8')
    st.download_button(
        "📥 Unduh Excel (CSV)",
        data=csv_bytes,
        file_name=f"Rekap_{selected_class}.csv",
        mime="text/csv",
    )

    pdf_like_text = rekap_df.to_string(index=False)
    st.download_button(
        "📄 Unduh PDF Sederhana",
        data=pdf_like_text.encode('utf-8'),
        file_name=f"Rekap_{selected_class}.pdf",
        mime="application/pdf",
    )

# =============================
# HALAMAN: DASHBOARD & LAPORAN (LEADERBOARD KELAS)
# =============================

def page_dashboard(df, selected_class):
    st.header("📊 Dashboard & Laporan Progres Kelas")

    if selected_class == "Pilih Kelas":
        st.info("Pilih kelas di sidebar untuk melihat dashboard.")
        return

    st.subheader(f"Papan Peringkat Kelas {selected_class}")

    df_current = ensure_columns(df.copy())
    leaderboard_df = (
        df_current[df_current['Kelas'] == selected_class]
        .sort_values(by='Total_Ayat_Lulus', ascending=False)
        .reset_index(drop=True)
    )
    leaderboard_df.index = leaderboard_df.index + 1

    display_cols = [
        'Nama_Murid',
        'NIS',
        'Kelas',
        'Total_Ayat_Lulus',
        'Update_Terakhir',
        'Guru_Pencatat',
        'ID_Murid',
    ]

    column_mapping = {
        'Nama_Murid': 'Murid',
        'NIS': 'NIS',
        'Kelas': 'Kelas',
        'Total_Ayat_Lulus': 'Total Ayat Lulus',
        'Update_Terakhir': 'Update Terakhir',
        'Guru_Pencatat': 'Dicatat Oleh',
        'ID_Murid': 'ID',
    }

    st.dataframe(
        leaderboard_df[display_cols].rename(columns=column_mapping),
        use_container_width=True,
    )

    st.markdown("---")
    st.subheader("Grafik Progres Ayat Lulus per Murid")

    chart = px.bar(
        leaderboard_df,
        x='Nama_Murid',
        y='Total_Ayat_Lulus',
        color='Total_Ayat_Lulus',
        title=f"Total Ayat Lulus Tiap Murid - {selected_class}",
    )
    st.plotly_chart(chart, use_container_width=True)

    st.markdown("---")
    st.subheader("Detail Progres Murid per Surah")

    for _, row in leaderboard_df.iterrows():
        with st.expander(
            f"⭐ {row['Nama_Murid']} - Total Lulus: {row['Total_Ayat_Lulus']} Ayat (Dicatat oleh {row.get('Guru_Pencatat','')})"
        ):
            status_dict = json.loads(row['Status_Hafalan'])
            for surah, ayat_list in status_dict.items():
                total_ayat_surah = JUZ_AMMA_MAP[surah]
                lulus_count = ayat_list.count(1)
                mengulang_count = ayat_list.count(2)
                belum_count = ayat_list.count(0)

                progress_ratio = (lulus_count / total_ayat_surah) if total_ayat_surah > 0 else 0
                st.progress(
                    progress_ratio,
                    text=(
                        f"{surah} | Lulus: {lulus_count}/{total_ayat_surah} | "
                        f"Mengulang: {mengulang_count} | Belum: {belum_count}"
                    ),
                )

# =============================
# HALAMAN BARU: 📜 RIWAYAT SETORAN
# =============================

def page_riwayat_setoran():
    st.header("📜 Riwayat Setoran Hafalan (Log Harian)")

    if not os.path.exists(LOG_FILE):
        st.info("Belum ada data log setoran.")
        return

    df_log = pd.read_csv(LOG_FILE)

    # Tambahkan kolom Tanggal (date-only) utk filter harian
    df_log['Tanggal'] = pd.to_datetime(df_log['Timestamp']).dt.date

    tanggal_unik = sorted(df_log['Tanggal'].unique(), reverse=True)
    guru_unik = ["Semua Guru"] + sorted(df_log['Guru_Pencatat'].dropna().unique())

    col1, col2 = st.columns(2)
    selected_date = col1.selectbox("Tanggal", ["Semua Tanggal"] + [str(t) for t in tanggal_unik])
    selected_guru = col2.selectbox("Guru Pencatat", guru_unik)

    df_filtered = df_log.copy()
    if selected_date != "Semua Tanggal":
        df_filtered = df_filtered[df_filtered['Tanggal'].astype(str) == selected_date]
    if selected_guru != "Semua Guru":
        df_filtered = df_filtered[df_filtered['Guru_Pencatat'] == selected_guru]

    st.dataframe(df_filtered, use_container_width=True)

    csv_bytes = df_filtered.to_csv(index=False).encode('utf-8')
    st.download_button(
        "📥 Unduh CSV Riwayat Terpilih",
        data=csv_bytes,
        file_name="riwayat_setoran.csv",
        mime="text/csv",
    )

# =============================
# HALAMAN BARU: 📅 LAPORAN BULANAN
# =============================

def page_laporan_bulanan():
    st.header("📅 Laporan Bulanan Hafalan Juz Amma")

    if not os.path.exists(DB_FILE):
        st.info("Belum ada data hafalan yang tersimpan.")
        return

    df = pd.read_csv(DB_FILE)

    if df.empty:
        st.warning("Database masih kosong.")
        return

    # --- Dapatkan bulan & tahun laporan sekarang ---
    import calendar
    today = datetime.now()
    bulan_text = calendar.month_name[today.month]
    tahun_text = today.year
    judul_laporan = f"Laporan Hafalan Juz Amma Bulan {bulan_text} {tahun_text}"

    st.info("Laporan ini menampilkan rekap per siswa, surah yang sudah dinyatakan *lulus*, dan persentase hafalan terhadap seluruh Juz Amma.")

    hasil_data = []
    for _, row in df.iterrows():
        try:
            status_dict = json.loads(row["Status_Hafalan"])
        except Exception:
            continue

        surah_lulus = []
        total_ayat_lulus = 0
        for surah, ayat_list in status_dict.items():
            total_ayat_lulus += ayat_list.count(1)
            if all(a == 1 for a in ayat_list) and len(ayat_list) > 0:
                surah_lulus.append(surah)

        persen = round((total_ayat_lulus / TOTAL_AYAT_JUZ_AMMA) * 100, 2)

        hasil_data.append({
            "NIS": row.get("NIS", ""),
            "Nama": row["Nama_Murid"],
            "Kelas": row["Kelas"],
            "Surah Lulus": ", ".join(surah_lulus) if surah_lulus else "-",
            "% Hafalan Juz Amma": persen
        })

    laporan_df = pd.DataFrame(hasil_data)
    st.dataframe(laporan_df, use_container_width=True)

    # LAPORAN TAHUNAN YTD

def page_laporan_tahunan():
    st.header("📆 Laporan Tahunan (Year-to-Date) Hafalan Juz Amma")

    if not os.path.exists(LOG_FILE) or not os.path.exists(DB_FILE):
        st.warning("Data hafalan belum lengkap (log atau database tidak ditemukan).")
        return

    df_log = pd.read_csv(LOG_FILE)
    df_data = pd.read_csv(DB_FILE)

    if df_log.empty or df_data.empty:
        st.warning("Belum ada data untuk ditampilkan.")
        return

    df_log["Timestamp"] = pd.to_datetime(df_log["Timestamp"], errors="coerce")
    df_log = df_log.dropna(subset=["Timestamp"])
    df_log["Tahun"] = df_log["Timestamp"].dt.year
    tahun_ini = datetime.now().year

    df_log_tahun = df_log[df_log["Tahun"] == tahun_ini].copy()
    if df_log_tahun.empty:
        st.info(f"Belum ada data setoran untuk tahun {tahun_ini}.")
        return

    hasil = []
    for _, murid in df_data.iterrows():
        murid_log = df_log_tahun[df_log_tahun["ID_Murid"] == murid["ID_Murid"]]
        if murid_log.empty:
            continue

        jumlah_setoran = len(murid_log)
        jumlah_lulus = (murid_log["Status"] == "Lulus").sum()
        jumlah_mengulang = (murid_log["Status"] == "Mengulang").sum()

        # total ayat lulus kumulatif
        df_lulus = murid_log[murid_log["Status"] == "Lulus"].copy()
        df_lulus["Jumlah_Ayat"] = df_lulus["Ayat_Sampai"] - df_lulus["Ayat_Dari"] + 1
        total_ayat_lulus = df_lulus["Jumlah_Ayat"].sum()
        persen = round((total_ayat_lulus / TOTAL_AYAT_JUZ_AMMA) * 100, 2)

        # ambil surah yang sudah lulus penuh dari data utama
        try:
            status_dict = json.loads(murid["Status_Hafalan"])
            surah_lulus = [
                s for s, a in status_dict.items() if all(x == 1 for x in a) and len(a) > 0
            ]
        except Exception:
            surah_lulus = []

        hasil.append({
            "NIS": murid.get("NIS", ""),
            "Nama": murid["Nama_Murid"],
            "Kelas": murid["Kelas"],
            "Jumlah Setoran Tahun Ini": jumlah_setoran,
            "Jumlah Lulus": jumlah_lulus,
            "Jumlah Mengulang": jumlah_mengulang,
            "% Hafalan Juz Amma": persen,
            "Surah Lulus": ", ".join(surah_lulus) if surah_lulus else "-",
        })

    laporan_df = pd.DataFrame(hasil)
    st.dataframe(laporan_df, use_container_width=True)

    # === Simpan Excel ===
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.cell.cell import MergedCell

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = f"Laporan {tahun_ini}"
        laporan_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        sheet = writer.sheets[sheet_name]

        sheet["A1"] = f"Laporan Tahunan Hafalan Juz Amma - Tahun {tahun_ini}"
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")

        max_col = len(laporan_df.columns)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

        # format header tabel
        for col_cells in sheet.iter_cols(min_row=3, max_row=3):
            for cell in col_cells:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        # lebar kolom otomatis
        for column_cells in sheet.columns:
            first_real_cell = next((cell for cell in column_cells if not isinstance(cell, MergedCell)), None)
            if first_real_cell is None:
                continue
            column_letter = first_real_cell.column_letter
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_letter].width = length + 3

    file_name = f"Laporan_Hafalan_Tahunan_{tahun_ini}.xlsx"
    st.download_button(
        label="📥 Unduh Laporan Tahunan (Excel)",
        data=output.getvalue(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # --------------------------
    # SIMPAN SEBAGAI EXCEL (XLSX)
    # --------------------------
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        laporan_df.to_excel(writer, sheet_name="Laporan Bulanan", index=False, startrow=2)

        sheet = writer.sheets["Laporan Bulanan"]

        # Tulis judul besar di baris 1
        sheet["A1"] = f"Laporan Tahunan Hafalan Juz Amma - Tahun {tahun_ini}"
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")

        # Merge sel judul agar lebar seperti tabel
        max_col = len(laporan_df.columns)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    # Lebarkan kolom otomatis — skip sel merge agar tidak error
    from openpyxl.cell.cell import MergedCell

    for column_cells in sheet.columns:
        # ambil huruf kolom dari sel pertama yang BUKAN merged
        first_real_cell = next((cell for cell in column_cells if not isinstance(cell, MergedCell)), None)
        if first_real_cell is None:
            continue
        column_letter = first_real_cell.column_letter
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_letter].width = length + 3

    file_name = f"Laporan_Hafalan_Tahunan_{tahun_ini}.xlsx"

    st.download_button(
        label="📥 Unduh Laporan Bulanan (Excel)",
        data=output.getvalue(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# =============================
# SIDEBAR (NAVIGASI + ADMINISTRASI)
# =============================

def sidebar_controls(df):
    st.sidebar.title("Navigasi")

    # --- tampilkan logo sekolah jika ada ---
    logo_path = os.path.join(BASE_DIR, "logo.png")
    if os.path.exists(logo_path):
        st.sidebar.image(logo_path, width=120)
    else:
        st.sidebar.markdown("**SMP Negeri 9 Banjar**")

    # --- menu utama aplikasi ---
    menu = st.sidebar.radio(
        "Pilih Tampilan",
        [
            "Pencatatan Hafalan",
            "Rekap Per Surah",
            "Dashboard & Laporan",
            "📜 Riwayat Setoran",
            "📅 Laporan Bulanan",
            "📆 Laporan Tahunan (YTD)",
            "👤 Profil Murid",
            "🏫 Pantauan Kelas",
        ],
    )

    # --- pilih guru dan kelas ---
    guru_list = load_guru_list()
    selected_guru = st.sidebar.selectbox("Nama Guru Pencatat", guru_list)

    kelas_list = ["Pilih Kelas"] + sorted(df["Kelas"].unique().tolist())
    selected_class = st.sidebar.selectbox("Kelas", kelas_list)

    # --- identitas sekolah di sidebar ---
    st.sidebar.markdown("---")
    st.sidebar.markdown(
        """
        **Aplikasi Hafalan Juz Amma**  
        _SMP Negeri 9 Banjar_  
        Pengembang: **Agus Sugiharto Sapari, S.Pd.**  
        © 2025
        """
    )

    # --- nilai yang dikembalikan ke fungsi utama ---
    return menu, selected_class, selected_guru


    # Pilih kelas
    kelas_list = ["Pilih Kelas"] + sorted(df['Kelas'].unique().tolist())
    selected_class = st.sidebar.selectbox("Kelas", kelas_list, key="kelas_filter")

    # ====================
    # ADMIN GURU (CRUD)
    # ====================
    st.sidebar.markdown("---")
    st.sidebar.title("🛠️ Administrasi Data Murid")
    st.sidebar.caption("Kelola data murid (tambah, impor, hapus).")

    # Tambah murid baru manual
    with st.sidebar.expander("➕ Tambah Murid Baru (Manual)"):
        with st.form("add_student_form"):
            new_name = st.text_input("Nama Lengkap Murid", max_chars=100)
            new_nis = st.text_input("Nomor Induk Siswa (NIS)", max_chars=20, value="")
            existing_classes = (
                sorted(df['Kelas'].unique().tolist()) if not df.empty else []
            )
            new_kelas = st.text_input(
                "Kelas (contoh: VII-A, VIII-B)",
                max_chars=10,
                value=existing_classes[0] if existing_classes else "VII-A",
            )
            add_submitted = st.form_submit_button("Simpan Murid Baru")
            if add_submitted:
                if new_name and new_kelas:
                    add_new_student(new_name, new_kelas, new_nis)
                    st.rerun()
                else:
                    st.error("Nama dan Kelas tidak boleh kosong.")

    # Impor massal CSV
    with st.sidebar.expander("⬆️ Impor Massal (CSV)"):
        st.markdown(
            "**Kolom wajib:** `Nama_Murid`, `Kelas`.\n\n"
            "**Opsional:** `NIS`.\n\n"
            "Gunakan pemisah `;` (titik koma)."
        )
        uploaded_file = st.file_uploader(
            "Pilih file CSV", type=["csv"], key="csv_uploader"
        )
        if uploaded_file is not None:
            if st.button("Proses Impor Data"):
                import_students_from_csv(uploaded_file)
                st.rerun()

    # Hapus murid permanen
    with st.sidebar.expander("🗑️ Hapus Murid"):
        st.warning("PERINGATAN: Penghapusan permanen. Tidak bisa dibatalkan.")

        delete_df = df.copy()
        existing_classes_delete = (
            sorted(delete_df['Kelas'].unique().tolist()) if not delete_df.empty else []
        )
        delete_class_filter = st.selectbox(
            "Filter Berdasarkan Kelas",
            ['Semua Kelas'] + existing_classes_delete,
            key="delete_class_filter",
        )

        filtered_delete_df = delete_df
        if delete_class_filter != 'Semua Kelas':
            filtered_delete_df = delete_df[delete_df['Kelas'] == delete_class_filter].copy()

        internal_delete_map = {}
        for _, row in filtered_delete_df.iterrows():
            internal_key = f"{row['Nama_Murid']} - Kelas: {row['Kelas']} |ID:{row['ID_Murid']}"
            internal_delete_map[internal_key] = row['ID_Murid']

        sorted_internal_keys = sorted(internal_delete_map.keys())
        display_list_delete = ['Pilih Murid yang Akan Dihapus'] + [
            key.rsplit(' |ID:', 1)[0] for key in sorted_internal_keys
        ]

        selected_display_string = st.selectbox(
            "Pilih Murid yang Akan Dihapus",
            display_list_delete,
            key="delete_student_select",
        )

        if selected_display_string != 'Pilih Murid yang Akan Dihapus':
            # Cocokkan lagi ke internal key
            start_of_internal_key = selected_display_string
            found_key = next(
                (
                    key for key in sorted_internal_keys
                    if key.startswith(start_of_internal_key + ' |ID:')
                ),
                None,
            )
            if found_key:
                student_id_to_delete = internal_delete_map[found_key]
                student_name_to_delete = selected_display_string.split(' - Kelas:')[0]
                st.error(
                    f"Anda yakin ingin menghapus **{student_name_to_delete}** (ID: {student_id_to_delete}) secara permanen?"
                )
                if st.button(
                    f"✅ KONFIRMASI HAPUS {student_name_to_delete}",
                    key="confirm_delete_button",
                ):
                    delete_student(st.session_state.df.copy(), student_id_to_delete, student_name_to_delete)
                    st.rerun()
            else:
                st.warning("Murid yang dipilih tidak dapat diidentifikasi. Coba filter ulang.")

    st.sidebar.markdown("---")
    st.sidebar.markdown(
        """
        **Aplikasi Hafalan Juz Amma**  
        _SMP Negeri 9 Banjar_  
        Pengembang: **Agus Sugiharto Sapari, S.Pd.**  
        © 2025
        """
    )

    
    return menu, selected_class, selected_guru

# =============================
# MAIN APP FLOW
# =============================

def main_app():
    df = ensure_columns(st.session_state.df.copy())

    menu, selected_class, selected_guru = sidebar_controls(df)

    if menu == "Pencatatan Hafalan":
        page_pencatatan_hafalan(df, selected_class, selected_guru)

    elif menu == "Rekap Per Surah":
        page_rekap_per_surah(df, selected_class)

    elif menu == "Dashboard & Laporan":
        page_dashboard(df, selected_class)

    elif menu == "📜 Riwayat Setoran":
        page_riwayat_setoran()

    elif menu == "📅 Laporan Bulanan":
        page_laporan_bulanan()
        
    elif menu == "📆 Laporan Tahunan (YTD)":
        page_laporan_tahunan()
    
    elif menu == "👤 Profil Murid":
        page_profil_murid(df)
        
    elif menu == "🏫 Pantauan Kelas":
        page_pantauan_kelas(df)
# =============================
# HALAMAN BARU: 👤 PROFIL MURID
# =============================

def page_profil_murid(df):
    st.header("👤 Profil Murid")

    if not os.path.exists(LOG_FILE):
        st.info("Belum ada data log setoran.")
        return

    kelas_list = sorted(df["Kelas"].unique().tolist())
    selected_class = st.selectbox("Pilih Kelas", kelas_list, key="profil_kelas")

    class_df = df[df["Kelas"] == selected_class]
    murid_map = {f"{r['Nama_Murid']} (ID:{r['ID_Murid']})": r["ID_Murid"] for _, r in class_df.iterrows()}
    selected_murid = st.selectbox("Pilih Murid", ["Pilih Murid"] + list(murid_map.keys()), key="profil_murid")

    if selected_murid == "Pilih Murid":
        return

    murid_id = murid_map[selected_murid]
    df_log = pd.read_csv(LOG_FILE)
    df_log["Timestamp"] = pd.to_datetime(df_log["Timestamp"])
    df_log["Tanggal"] = df_log["Timestamp"].dt.date
    df_murid = df_log[df_log["ID_Murid"] == murid_id]

    if df_murid.empty:
        st.info("Belum ada histori setoran untuk murid ini.")
        return

    total_setoran = len(df_murid)
    total_lulus = (df_murid["Status"] == "Lulus").sum()
    total_mengulang = (df_murid["Status"] == "Mengulang").sum()
    col1, col2, col3 = st.columns(3)
    col1.metric("Total Setoran", total_setoran)
    col2.metric("Lulus", total_lulus)
    col3.metric("Mengulang", total_mengulang)

    # Hanya ayat Lulus (kumulatif)
    df_lulus = df_murid[df_murid["Status"] == "Lulus"].copy()
    df_lulus["Jumlah_Ayat"] = df_lulus["Ayat_Sampai"] - df_lulus["Ayat_Dari"] + 1
    progres = df_lulus.groupby("Tanggal")["Jumlah_Ayat"].sum().reset_index()
    progres["Kumulatif"] = progres["Jumlah_Ayat"].cumsum()

    st.subheader("📈 Grafik Perkembangan Hafalan (Ayat Lulus Kumulatif)")
    if not progres.empty:
        fig = px.line(progres, x="Tanggal", y="Kumulatif", markers=True, title="Grafik Kumulatif Ayat Lulus")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Belum ada data 'Lulus' untuk murid ini.")

    st.subheader("📚 Surah yang Paling Sering Disetorkan")
    surah_count = df_murid.groupby(["Surah", "Status"]).size().reset_index(name="Jumlah_Setoran")
    fig2 = px.bar(surah_count, x="Surah", y="Jumlah_Setoran", color="Status", barmode="group")
    st.plotly_chart(fig2, use_container_width=True)


# =============================
# HALAMAN BARU: 🏫 PANTAUAN KELAS
# =============================

def page_pantauan_kelas(df):
    st.header("🏫 Pantauan Per Kelas")

    if not os.path.exists(LOG_FILE):
        st.info("Belum ada data log setoran.")
        return

    df_log = pd.read_csv(LOG_FILE)
    df_log["Timestamp"] = pd.to_datetime(df_log["Timestamp"])
    df_log["Tanggal"] = df_log["Timestamp"].dt.date
    df_log["Jumlah_Ayat"] = df_log["Ayat_Sampai"] - df_log["Ayat_Dari"] + 1
    df_lulus = df_log[df_log["Status"] == "Lulus"]

    kelas_list = sorted(df["Kelas"].unique().tolist())
    selected_class = st.selectbox("Pilih Kelas", kelas_list, key="pantau_kelas")
    df_kelas = df_lulus[df_lulus["Kelas"] == selected_class]

    if df_kelas.empty:
        st.info("Belum ada data 'Lulus' untuk kelas ini.")
        return

    progres = df_kelas.groupby("Tanggal")["Jumlah_Ayat"].sum().reset_index()
    progres["Kumulatif"] = progres["Jumlah_Ayat"].cumsum()
    st.subheader(f"📈 Perkembangan Kelas {selected_class}")
    fig = px.line(progres, x="Tanggal", y="Kumulatif", markers=True, title=f"Total Ayat Lulus Kelas {selected_class}")
    st.plotly_chart(fig, use_container_width=True)

    st.subheader("👩‍🏫 Guru Pencatat Teraktif")
    guru_rank = (
        df_kelas.groupby("Guru_Pencatat")
        .size()
        .reset_index(name="Jumlah_Setoran_Lulus")
        .sort_values("Jumlah_Setoran_Lulus", ascending=False)
    )
    st.dataframe(guru_rank, use_container_width=True)
    fig2 = px.bar(guru_rank, x="Guru_Pencatat", y="Jumlah_Setoran_Lulus", title=f"Aktivitas Guru di {selected_class}")
    st.plotly_chart(fig2, use_container_width=True)

# =============================
# ENTRY POINT
# =============================

def show_footer():
    st.markdown(
        """
        <hr style="margin-top: 40px; margin-bottom: 10px;">

        <div style="text-align: center; font-size: 14px; color: #555;">
            <strong>Aplikasi Catatan Hafalan Juz Amma</strong><br>
            SMP Negeri 9 Banjar<br>
            <em>Dikembangkan oleh:</em> Agus Sugiharto Sapari, S.Pd.<br>
            © 2025 SMP Negeri 9 Banjar. Seluruh hak cipta dilindungi.
        </div>
        """,
        unsafe_allow_html=True,
    )

if __name__ == "__main__":
    if not os.path.exists(DB_FILE):
        initialize_database(DB_FILE)
    main_app()
    show_footer()





















